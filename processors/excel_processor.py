#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel処理モジュール
"""
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from loguru import logger

from utils import backup_file, extract_employee_name_from_filename


def write_to_excel(template_path: str, output_path: str, df: pd.DataFrame, csv_filename: str):
    """
    ひな型Excelに勤怠データを書き込む

    Args:
        template_path: テンプレートExcelファイルのパス
        output_path: 出力先パス
        df: 書き込むデータ
        csv_filename: 元のCSVファイル名

    Raises:
        FileNotFoundError: テンプレートファイルが存在しない場合
        ValueError: データやテンプレートの形式が不正な場合
    """
    try:
        # テンプレートの存在確認
        if not os.path.exists(template_path):
            logger.error(f"テンプレートファイルが存在しません: {template_path}")
            raise FileNotFoundError(f"テンプレートファイルが存在しません: {template_path}")

        # 出力ディレクトリの確認
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"出力ディレクトリを作成しました: {output_dir}")

        # 出力ファイルが既に存在する場合はバックアップを作成
        if os.path.exists(output_path):
            backup_path = backup_file(output_path)
            logger.info(f"既存ファイルをバックアップしました: {backup_path}")

        # テンプレートを読み込み
        logger.info(f"テンプレートを読み込んでいます: {template_path}")
        wb = openpyxl.load_workbook(template_path)

        # 勤務表シートが存在するか確認
        if "勤務表" not in wb.sheetnames:
            logger.error("テンプレートに「勤務表」シートがありません")
            raise ValueError("テンプレートに「勤務表」シートがありません")

        sheet = wb["勤務表"]

        # 従業員名をCSVファイル名から取得
        employee_name = extract_employee_name_from_filename(os.path.basename(csv_filename))
        if employee_name:
            logger.info(f"従業員名を検出しました: {employee_name}")
            sheet["G1"] = employee_name
        else:
            logger.warning("ファイル名から従業員名を検出できませんでした")

        # CSVから月情報を取得
        if "日付" not in df.columns:
            logger.error("データに日付カラムがありません")
            raise ValueError("データに日付カラムがありません")

        if len(df) == 0:
            logger.error("データが空です")
            raise ValueError("データが空です")

        # 年月を取得
        try:
            first_date = df["日付"].min()
            month_value = first_date.month
            year_value = first_date.year

            # シートに年月を設定
            sheet["F5"] = year_value
            sheet["H5"] = month_value

            logger.info(f"年月を設定しました: {year_value}年{month_value}月")
        except Exception as e:
            logger.error(f"年月の取得に失敗しました: {str(e)}")
            raise ValueError(f"年月の取得に失敗しました: {str(e)}")

        # 勤怠データを書き込み
        logger.info("勤怠データを書き込んでいます...")

        # 書き込み開始行
        start_row = 11

        # データの件数によって処理
        if len(df) > 31:
            logger.warning(f"データが31日分を超えています: {len(df)}行, 最初の31行のみを処理します")
            df = df.sort_values("日付").iloc[:31]

        # 日数に合わせて行を調整（必要に応じて）
        date_range = pd.date_range(start=f"{year_value}-{month_value}-01", periods=31, freq='D')

        for i, date in enumerate(date_range, start=0):
            row_idx = start_row + i
            current_day = i + 1

            # その日のデータを抽出
            day_data = df[df["日付"].dt.day == current_day]

            # 日付列に日付数式を設定
            sheet[f"A{row_idx}"] = f"=DATE({year_value},{month_value},{current_day})"

            # その日のデータがある場合のみ書き込み
            if not day_data.empty:
                row = day_data.iloc[0]

                # 始業時刻
                if "始業時刻" in row and pd.notnull(row["始業時刻"]):
                    sheet[f"C{row_idx}"] = row["始業時刻"]

                # 終業時刻
                if "終業時刻" in row and pd.notnull(row["終業時刻"]):
                    sheet[f"D{row_idx}"] = row["終業時刻"]

                # 休憩時間
                is_workday = row.get("勤怠種別") not in ["未入力", "所定休日", "法定休日"]
                sheet[f"E{row_idx}"] = "1:00" if is_workday else ""

                # 総勤務時間
                if "総勤務時間" in row and pd.notnull(row["総勤務時間"]):
                    sheet[f"F{row_idx}"] = row["総勤務時間"]

                # その他のカラムがあれば追加
                if "時間外労働" in row and pd.notnull(row["時間外労働"]):
                    sheet[f"G{row_idx}"] = row["時間外労働"]

                if "深夜労働" in row and pd.notnull(row["深夜労働"]):
                    sheet[f"H{row_idx}"] = row["深夜労働"]

        # ファイルの保存
        wb.save(output_path)
        logger.info(f"Excelファイルを保存しました: {output_path}")

        return True

    except Exception as e:
        logger.exception(f"Excelファイルの書き込みに失敗しました: {str(e)}")
        raise


def read_excel_template(template_path: str):
    """
    Excelテンプレートを読み込み、シート名や列の構造などを取得

    Args:
        template_path: テンプレートファイルのパス

    Returns:
        dict: テンプレートの構造情報
    """
    try:
        # テンプレートの存在確認
        if not os.path.exists(template_path):
            logger.error(f"テンプレートファイルが存在しません: {template_path}")
            raise FileNotFoundError(f"テンプレートファイルが存在しません: {template_path}")

        # テンプレートを読み込み
        wb = openpyxl.load_workbook(template_path)

        # 情報を収集
        template_info = {
            "シート名": wb.sheetnames,
            "シート構造": {}
        }

        # 各シートの構造を取得
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 使用されている範囲を取得
            min_row, max_row = sheet.min_row, sheet.max_row
            min_col, max_col = sheet.min_column, sheet.max_column

            # 固定セル値を取得
            fixed_cells = {}
            for row in range(min_row, min(max_row + 1, 10)):  # 最初の10行までを調査
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value:
                        fixed_cells[f"{get_column_letter(col)}{row}"] = str(cell.value)

            # シートの構造情報
            template_info["シート構造"][sheet_name] = {
                "使用範囲": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "固定セル": fixed_cells
            }

        logger.info(f"テンプレート構造を解析しました: {template_path}")
        return template_info

    except Exception as e:
        logger.exception(f"テンプレートの解析に失敗しました: {str(e)}")
        raise